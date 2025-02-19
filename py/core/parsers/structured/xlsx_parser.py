# type: ignore
from io import BytesIO
from typing import AsyncGenerator

import networkx as nx
import numpy as np
from openpyxl import load_workbook

from core.base.parsers.base_parser import AsyncParser
from core.base.providers import (
    CompletionProvider,
    DatabaseProvider,
    IngestionConfig,
)


class XLSXParser(AsyncParser[str | bytes]):
    """A parser for XLSX data."""

    def __init__(
        self,
        config: IngestionConfig,
        database_provider: DatabaseProvider,
        llm_provider: CompletionProvider,
    ):
        self.database_provider = database_provider
        self.llm_provider = llm_provider
        self.config = config
        self.load_workbook = load_workbook

    async def ingest(
        self, data: bytes, *args, **kwargs
    ) -> AsyncGenerator[str, None]:
        """Ingest XLSX data and yield text from each row."""
        if isinstance(data, str):
            raise ValueError("XLSX data must be in bytes format.")

        wb = self.load_workbook(filename=BytesIO(data))
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                yield ", ".join(map(str, row))


class XLSXParserAdvanced(AsyncParser[str | bytes]):
    """A parser for XLSX data."""

    # identifies connected components in the excel graph and extracts data from each component
    def __init__(
        self, config: IngestionConfig, llm_provider: CompletionProvider
    ):
        self.llm_provider = llm_provider
        self.config = config
        self.nx = nx
        self.np = np
        self.load_workbook = load_workbook

    def connected_components(self, arr):
        g = self.nx.grid_2d_graph(len(arr), len(arr[0]))
        empty_cell_indices = list(
            zip(*self.np.where(arr is None), strict=False)
        )
        g.remove_nodes_from(empty_cell_indices)
        components = self.nx.connected_components(g)
        for component in components:
            rows, cols = zip(*component, strict=False)
            min_row, max_row = min(rows), max(rows)
            min_col, max_col = min(cols), max(cols)
            yield arr[min_row : max_row + 1, min_col : max_col + 1].astype(
                "str"
            )

    async def ingest(
        self, data: bytes, num_col_times_num_rows: int = 100, *args, **kwargs
    ) -> AsyncGenerator[str, None]:
        """Ingest XLSX data and yield text from each connected component."""
        if isinstance(data, str):
            raise ValueError("XLSX data must be in bytes format.")

        workbook = self.load_workbook(filename=BytesIO(data))

        for ws in workbook.worksheets:
            ws_data = self.np.array(
                [[cell.value for cell in row] for row in ws.iter_rows()]
            )
            for table in self.connected_components(ws_data):
                # parse like a csv parser, assumes that the first row has column names
                if len(table) <= 1:
                    continue

                num_rows = len(table)
                num_rows_per_chunk = num_col_times_num_rows // num_rows
                headers = ", ".join(table[0])
                # add header to each one
                for i in range(1, num_rows, num_rows_per_chunk):
                    chunk = table[i : i + num_rows_per_chunk]
                    yield (
                        headers
                        + "\n"
                        + "\n".join([", ".join(row) for row in chunk])
                    )
