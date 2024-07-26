from io import BytesIO
from typing import AsyncGenerator

from r2r.base.abstractions.document import DataType
from r2r.base.parsers.base_parser import AsyncParser


class XLSXParser(AsyncParser[DataType]):
    """A parser for XLSX data."""

    def __init__(self):
        try:
            from openpyxl import load_workbook

            self.load_workbook = load_workbook
        except ImportError:
            raise ValueError(
                "Error, `openpyxl` is required to run `XLSXParser`. Please install it using `pip install openpyxl`."
            )

    async def ingest(self, data: bytes) -> AsyncGenerator[str, None]:
        """Ingest XLSX data and yield text from each row."""
        if isinstance(data, str):
            raise ValueError("XLSX data must be in bytes format.")

        wb = self.load_workbook(filename=BytesIO(data))
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                yield ", ".join(map(str, row))


class XLSXParserAdvanced(AsyncParser[DataType]):
    """A parser for XLSX data."""

    # identifies connected components in the excel graph and extracts data from each component
    def __init__(self):
        try:
            import networkx as nx
            import numpy as np
            from openpyxl import load_workbook

            self.nx = nx
            self.np = np
            self.load_workbook = load_workbook

        except ImportError:
            raise ValueError(
                "Error, `networkx` and `numpy` are required to run `XLSXParserAdvanced`. Please install them using `pip install networkx numpy`."
            )

    def connected_components(self, arr):
        g = self.nx.grid_2d_graph(len(arr), len(arr[0]))
        empty_cell_indices = list(zip(*self.np.where(arr == None)))
        g.remove_nodes_from(empty_cell_indices)
        components = self.nx.connected_components(g)
        for component in components:
            rows, cols = zip(*component)
            min_row, max_row = min(rows), max(rows)
            min_col, max_col = min(cols), max(cols)
            yield arr[min_row : max_row + 1, min_col : max_col + 1].astype(
                "str"
            )

    async def ingest(
        self, data: bytes, num_col_times_num_rows: int = 100
    ) -> AsyncGenerator[str, None]:
        """Ingest XLSX data and yield text from each connected component."""
        if isinstance(data, str):
            raise ValueError("XLSX data must be in bytes format.")

        workbook = self.load_workbook(filename=BytesIO(data))

        for ws in workbook.worksheets:
            ws_data = self.np.array(
                [[cell.value for cell in row] for row in ws.iter_rows()]
            )
            for table in self.connected_components(ws_data):

                # parse like a csv parser, assumes that the first row has column names
                if len(table) <= 1:
                    continue

                num_cols, num_rows = len(table[0]), len(table)
                num_rows_per_chunk = num_col_times_num_rows // num_rows
                headers = ", ".join(table[0])
                # add header to each one
                for i in range(1, num_rows, num_rows_per_chunk):
                    chunk = table[i : i + num_rows_per_chunk]
                    yield headers + "\n" + "\n".join(
                        [", ".join(row) for row in chunk]
                    )


# async def main():
#     csv_file = '/Users/shreyas/parse_this.xlsx'
#     parser = XLSXParserAdvanced()

#     with open(csv_file, 'rb') as file:
#         file_content = file.read()

#     async for chunk in parser.ingest(BytesIO(file_content)):
#         print("Chunk:")
#         print(chunk)
#         print("---")  # Separator between chunks

#     import pdb; pdb.set_trace()


# if __name__== '__main__':
#     import asyncio
#     out = asyncio.run(main())
